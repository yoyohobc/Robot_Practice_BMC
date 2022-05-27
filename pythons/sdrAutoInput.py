from openpyxl import load_workbook
from robot.libraries.BuiltIn import BuiltIn
import json
curpath = BuiltIn().get_variable_value("${PYTHONS_PATH}") + '/'
resPath = BuiltIn().get_variable_value("${RESULT_PATH}") + '/'
outputXlsx = resPath + 'sdrAutoInput.xlsx'
#XLSX_MODEL
wb = load_workbook(BuiltIn().get_variable_value("${XLSX_MODEL}"))
ws = wb["SDR Table"]
thrkeys = ["Lower Critical","Lower Non-Critical","Upper Non-Critical","Upper Critical"]
def sdr_input_xlsx():
    f = open(curpath + 'sdrTable.json')
    f1 = open(curpath + 'spec.json')
    sdrTable = json.load(f)
    specTable = json.load(f1)
    for i in range(1,420):
        sensorName = ws['B'+str(i)].value
        errorResult1 = ''
        errorResult2 = ''
        sdr = sdrTable.get(sensorName)
        spec = specTable.get(sensorName)
        if sdr == None and spec != None:
            ws['G'+str(i+1)] = 'F'
            ws['G'+str(i+2)] = 'F'
            ws['L'+str(i+1)] = "ipmitool sdr list doesn't get this"
            ws['L'+str(i+2)] = "ipmitool sensor get <sensor ID> doesn't get this"
        if sdr != None and spec != None:
            for specKey,specValue in spec.items():
                testValue = sdr.get(specKey)
                if specValue!=0 and isinstance(specValue,float) and isinstance(testValue,float):
                    percent = abs((testValue - specValue)/specValue)
                    if percent < 0.05:
                        continue
                if specValue != testValue:
                    eRes = specKey + '\n' + 'Spec:'+ str(specValue) + ' Test:'+ str(testValue) + '\n'
                    if specKey in thrkeys:
                        errorResult2 += eRes
                    else:
                        errorResult1 += eRes
            ws['G'+str(i+1)] = 'P' if errorResult1 == '' else 'F'
            ws['G'+str(i+2)] = 'P' if errorResult2 == '' else 'F'
            ws['L'+str(i+1)] = errorResult1 + sdr.get('sdrResult')
            ws['L'+str(i+2)] = errorResult2 + sdr.get('sdrListResult')

    wb.save(outputXlsx)