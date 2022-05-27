from openpyxl import load_workbook
from robot.libraries.BuiltIn import BuiltIn
import json,os
curpath = BuiltIn().get_variable_value("${PYTHONS_PATH}") + '/'
resPath = BuiltIn().get_variable_value("${RESULT_PATH}") + '/'
redfish_ipv6 = BuiltIn().get_variable_value("${REDFISH_IPV6}")
scriptPath = BuiltIn().get_variable_value("${SCRIPTS_PATH}") + '/'
sh_path = scriptPath + redfish_ipv6 + '.sh'
outputXlsx = resPath + 'redfishAutoInput.xlsx'
methods = ['post','delete','patch','DELETE']

def replace_ip():
    with open(sh_path) as f:
        shFile = f.read()
    shIpv6 = shFile[shFile.index('[')+1:shFile.index(']')]
    redfishHost = BuiltIn().get_variable_value("${Redfish_Host}")
    if shIpv6 == redfishHost:
        return None
    else:
        shFile = shFile.replace(shIpv6,redfishHost)
        with open(sh_path, 'w+') as file:
            file.write(shFile)

def run_redfish_script():
    os.system(BuiltIn().get_variable_value("${CHMOD_CMD}")+' '+sh_path)  
    cmd_buf = sh_path + ' > ' + resPath + redfish_ipv6 + '.txt'
    os.system(cmd_buf)

def checkListInString(lists,strigs):
    for l in lists:
        if l in strigs:
            return True
    return False

def dumpJson(jsonName,dictVar):
    with open(curpath+jsonName, "w+") as f:
        json.dump(dictVar, f, indent = 4)

def redfishMapping():
    f = open(resPath+redfish_ipv6+'.txt')
    redfishLines = f.readlines()
    redfishMap = {}
    for i in range(len(redfishLines)):
        if 'curl' in redfishLines[i]:
            if checkListInString(methods,redfishLines[i]):
                continue
            else:
                if '/redfish/' in redfishLines[i]:
                    rKey = redfishLines[i][redfishLines[i].index('/redfish/'):-1]
                    logResult = ''
                    c = 0
                    for k in range(i,len(redfishLines)):
                        if 'curl' in redfishLines[k]:
                            c+=1
                            if c == 2:
                                break
                        logResult += redfishLines[k]

                    redfishMap[rKey] = {'log':logResult,'methods':'GET'}
    return redfishMap
def redfish_input():
    wb = load_workbook(BuiltIn().get_variable_value("${XLSX_MODEL}"))
    ws = wb["Redfish check"]
    redfishMap = redfishMapping()
    dumpJson('redfish.json',redfishMap)
    for i in range(1,140):
        method = ws['C'+str(i)].value
        redfishAPI = ws['D'+str(i)].value
        #print(method,redfishAPI)
        redfish = redfishMap.get(redfishAPI)
        if redfish != None and redfish.get('methods') == method:
            ws['H'+str(i)] = redfish.get('log')
    wb.save(outputXlsx)