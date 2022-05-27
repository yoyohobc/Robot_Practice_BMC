from openpyxl import load_workbook
import json 

wb = load_workbook('Astoria-SIT-TestPlan_BMCQA_EVT2.xlsx')
ws = wb["Redfish check"]
f = open('Redfish_check.txt')
redfishLines = f.readlines()
methods = ['post','delete','patch','DELETE']

def checkListInString(lists,strigs):
    for l in lists:
        if l in strigs:
            return True
    return False

def dumpJson(jsonName,dictVar):
    with open(jsonName, "w+") as f:
        json.dump(dictVar, f, indent = 4)

def redfishMapping():
    redfishMap = {}
    for i in range(len(redfishLines)):
        if 'curl' in redfishLines[i]:
            if checkListInString(methods,redfishLines[i]):
                continue
            else:
                if '/redfish/' in redfishLines[i]:
                    rKey = redfishLines[i][redfishLines[i].index('/redfish/'):-1]
                    logResult = ''
                    c = 0
                    for k in range(i,len(redfishLines)):
                        if 'curl' in redfishLines[k]:
                            c+=1
                            if c == 2:
                                break
                        logResult += redfishLines[k]

                    redfishMap[rKey] = {'log':logResult,'methods':'GET'}
    return redfishMap

redfishMap = redfishMapping()
dumpJson('redfish.json',redfishMap)
for i in range(1,140):
    method = ws['C'+str(i)].value
    redfishAPI = ws['D'+str(i)].value
    #print(method,redfishAPI)
    redfish = redfishMap.get(redfishAPI)
    if redfish != None and redfish.get('methods') == method:
        ws['H'+str(i)] = redfish.get('log')
wb.save('test.xlsx')